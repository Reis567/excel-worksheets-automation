import PySimpleGUI as sg
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment


class TelaPython:
    def __init__(self):
        # LAYOUT
        sg.theme('DarkPurple')
        layout = [
            [sg.Text('Nome:', size=(5, 0)), sg.Input(
                size=(15, 0), key='nome')],
            [sg.Text('Idade:', size=(5, 0)), sg.Input(
                size=(15, 0), key='idade')],
            [sg.Text("Quais provedores de email possui?")],
            [sg.Checkbox('Gmail', key='gmail'), sg.Checkbox(
                'Outlook', key='outlook'), sg.Checkbox('Yahoo', key='yahoo')],
            [sg.Text('Aceita cartão?')],
            [sg.Radio('Sim', 'cartoes', key='aceitaCartao'), sg.Radio(
                'Não', 'cartoes', key='naoAceitaCartao')],
            [sg.Button('Enviar', size=(25, 0))],
            [sg.Output(size=(30, 20), key='output')]
        ]
        # JANELA
        self.janela = sg.Window('Dados do Usuario', layout, element_justification='center', size=(
            1000, 600), resizable=True).finalize()

    def Iniciar(self):
        # Verifica se já existe um arquivo com dados salvos
        try:
            workbook = load_workbook('cadastro.xlsx')
            planilha_clientes = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            planilha_clientes = workbook.active

        # LOOP PARA REPETIR O CODIGO APOS ADICIONAR DADOS UMA VEZ
        while True:
            # EXTRAIR DADOS DA TELA
            self.button, self.values = self.janela.Read()

            # TRATAMENTO DE ERROS:
            # Verifica se a janela foi fechada
            if self.button is None:
                break
            # Verifica se todos os campos obrigatórios foram preenchidos
            if not self.values['nome'] or not self.values['idade']:
                sg.Popup('Por favor, preencha todos os campos obrigatórios.')
                continue
            # Verifica se a idade é um número
            if not self.values['idade'].isdigit():
                sg.Popup('A idade deve ser um número!')
                continue

            # Verifica se o nome é composto de letras
            # Faz com que nome e sobrenome não de bug pelo espaço entre eles
            nome_completo = self.values['nome']
            palavras_nome = nome_completo.split()

            for palavra in palavras_nome:
                if not palavra.isalpha():
                    sg.Popup('Nome inválido! Insira apenas letras.')
                    continue

            # TRATAMENTO DE ERROS FINALIZADO

            # Declaração de variaveis e prints
            nome = self.values['nome']
            idade = self.values['idade']
            aceita_gmail = self.values['gmail']
            aceita_outlook = self.values['outlook']
            aceita_yahoo = self.values['yahoo']
            aceita_cartao = self.values['aceitaCartao']
            nao_aceita_cartao = self.values['naoAceitaCartao']
            print(f'Nome:{nome}')
            print(f'Idade:{idade}')
            print(f'Aceita gmail:{aceita_gmail}')
            print(f'Aceita outlook:{aceita_outlook}')
            print(f'Aceita yahoo:{aceita_yahoo}')
            print(f'Aceita cartão:{aceita_cartao}')
            print(f'Não Aceita cartão:{nao_aceita_cartao}')

            # Adiciona as informações recebidas em uma nova linha da planilha
            # Alem disso coloquei para transformar o True e False em sim e não
            linha = [self.values['nome'], self.values['idade']]
            if self.values['gmail']:
                linha.append('sim')
            else:
                linha.append('não')
            if self.values['outlook']:
                linha.append('sim')
            else:
                linha.append('não')
            if self.values['yahoo']:
                linha.append('sim')
            else:
                linha.append('não')
            if self.values['aceitaCartao']:
                linha.append('sim')
            else:
                linha.append('não')
            if self.values['naoAceitaCartao']:
                linha.append('sim')
            else:
                linha.append('não')

            planilha_clientes.append(linha)

            # Alinha a linha adicionada ao centro
            for cell in planilha_clientes[planilha_clientes.max_row]:
                cell.alignment = Alignment(horizontal='center')

            # Salva as alterações no arquivo
            workbook.save('cadastro.xlsx')

            # Exibe uma mensagem informando que os dados foram salvos
            sg.Popup('Dados salvos com sucesso!')

            # Limpa o output
            self.janela.FindElement('output').update('')


tela = TelaPython()
tela.Iniciar()
